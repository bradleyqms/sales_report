"""
Base report generator class with shared functionality.

This module provides an abstract base class that encapsulates common report
generation patterns including config loading, date preparation, PDF styling,
table formatting, and multi-format exports.
"""

from abc import ABC, abstractmethod
import json
import datetime
import calendar
import pandas as pd
import logging
from pathlib import Path
from typing import Optional, Dict, List, Tuple

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from utils import get_current_year, get_prior_year, get_current_month, format_mtd_date_range

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class BaseReportGenerator(ABC):
    """
    Abstract base class for report generators.
    
    Provides common functionality for:
    - Config file loading and validation
    - Date calculations (current year, prior year, current month)
    - PDF document styling and table formatting
    - Multi-format exports (CSV, TXT, HTML, PDF, XLSX)
    
    Subclasses must implement:
    - calculate_report(): Generate report data as DataFrame
    - render_report(df): Display report to console
    """
    
    def __init__(self, config_path: str, sales_path: str, budget_path: str, prior_path: str,
                 report_date: Optional[datetime.datetime] = None):
        """
        Initialize the report generator.

        Args:
            config_path: Path to JSON configuration file
            sales_path: Path to sales data CSV
            budget_path: Path to budget data CSV
            prior_path: Path to prior year data CSV
            report_date: Optional business-date anchor derived from SAP Extract_Date.
                         When provided, all date logic (headers, budget/prior filtering)
                         derives from this value instead of the system clock.
        """
        self._report_date = report_date
        self.config = self._load_config(config_path)
        self._load_data_files(sales_path, budget_path, prior_path)
        self._prepare_dates()
    
    def _load_config(self, path: str) -> dict:
        """
        Load and validate JSON configuration file.
        
        Args:
            path: Path to config file
            
        Returns:
            Parsed configuration dictionary
            
        Raises:
            FileNotFoundError: If config file doesn't exist
            json.JSONDecodeError: If config file is invalid JSON
        """
        try:
            with open(path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logging.error(f"Config file not found: {path}")
            raise
        except json.JSONDecodeError as e:
            logging.error(f"Invalid JSON in config file: {e}")
            raise
    
    def _load_data_files(self, sales_path: str, budget_path: str, prior_path: str) -> None:
        """
        Load CSV data files into DataFrames.
        
        Args:
            sales_path: Path to sales data CSV
            budget_path: Path to budget data CSV
            prior_path: Path to prior year data CSV
            
        Raises:
            FileNotFoundError: If any data file doesn't exist
            pd.errors.EmptyDataError: If any data file is empty
        """
        try:
            self.df = pd.read_csv(sales_path)
            self.budget_df = pd.read_csv(budget_path)
            self.prior_df = pd.read_csv(prior_path)
        except FileNotFoundError as e:
            logging.error(f"Required data file not found: {e}")
            raise
        except pd.errors.EmptyDataError as e:
            logging.error(f"Data file is empty: {e}")
            raise
    
    def _prepare_dates(self) -> None:
        """
        Calculate and store commonly used date values.

        Sets instance variables:
        - now: Report datetime anchor (from report_date if provided, else current datetime)
        - current_year: Year from anchor (int)
        - prior_year: Prior year from anchor (int)
        - current_month: Month from anchor (int, 1-12)
        """
        self.now = self._report_date or datetime.datetime.now()
        self.current_year = get_current_year(self.now)
        self.prior_year = get_prior_year(self.now)
        self.current_month = get_current_month(self.now)
    
    def _filter_budget_for_month(self, date_format: str = '%d/%m/%Y') -> pd.DataFrame:
        """
        Filter budget data for the current month.
        
        Args:
            date_format: Date format string for parsing (default DD/MM/YYYY)
            
        Returns:
            DataFrame filtered to current month's budget data
        """
        budget_df = self.budget_df.copy()
        budget_df['Date'] = pd.to_datetime(budget_df['Date'], format=date_format, errors='coerce')
        budget_month = budget_df[budget_df['Date'].dt.month == self.current_month].copy()
        
        # Ensure numeric columns are parsed correctly
        if 'Value_kEUR' in budget_month.columns:
            budget_month['Value_kEUR'] = pd.to_numeric(budget_month['Value_kEUR'], errors='coerce').fillna(0)
        if 'Value_kUSD' in budget_month.columns:
            budget_month['Value_kUSD'] = pd.to_numeric(budget_month['Value_kUSD'], errors='coerce').fillna(0)
        
        return budget_month
    
    def _filter_prior_for_month(self, date_format: str = '%d/%m/%Y') -> pd.DataFrame:
        """
        Filter prior year data for the same month last year.
        
        Args:
            date_format: Date format string for parsing (default DD/MM/YYYY)
            
        Returns:
            DataFrame filtered to prior year's same month data
        """
        prior_df = self.prior_df.copy()
        prior_df['Date'] = pd.to_datetime(prior_df['Date'], format=date_format, errors='coerce')
        prior_month = prior_df[
            (prior_df['Date'].dt.year == self.prior_year) & 
            (prior_df['Date'].dt.month == self.current_month)
        ].copy()
        
        # Ensure numeric columns are parsed correctly
        if 'Value_kEUR' in prior_month.columns:
            prior_month['Value_kEUR'] = pd.to_numeric(prior_month['Value_kEUR'], errors='coerce').fillna(0)
        if 'Value_kUSD' in prior_month.columns:
            prior_month['Value_kUSD'] = pd.to_numeric(prior_month['Value_kUSD'], errors='coerce').fillna(0)
        
        return prior_month
    
    def _convert_to_keur(self, value_col: str = None) -> None:
        """
        Convert sales values to kEUR and store in 'kEUR' column.
        
        Args:
            value_col: Column name containing EUR values. If None, auto-detects.
        """
        if value_col is None:
            value_col = 'Value_in_EUR_converted' if 'Value_in_EUR_converted' in self.df.columns else 'Total Value (EUR)'
        
        self.df['kEUR'] = pd.to_numeric(self.df[value_col], errors='coerce').fillna(0) / 1000
    
    def get_pdf_styles(self) -> Dict[str, TableStyle]:
        """
        Get standard PDF table styles.
        
        Returns:
            Dictionary of TableStyle objects for different table types:
            - 'header': Styling for header row
            - 'data': Styling for data rows
            - 'total': Styling for total rows
            - 'grand_total': Styling for grand total rows
        """
        styles = {
            'header': TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Left align first column
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]),
            'data': TableStyle([
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ]),
            'total': TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]),
            'grand_total': TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightsteelblue),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ])
        }
        return styles
    
    def create_pdf_table(self, data: List[List], title: str, pagesize=A4) -> SimpleDocTemplate:
        """
        Create a styled PDF table document.
        
        Args:
            data: 2D list of table data (rows x columns)
            title: Document title
            pagesize: ReportLab page size (default A4)
            
        Returns:
            Configured SimpleDocTemplate ready to build
        """
        # This method would be implemented by subclasses with specific table layouts
        pass
    
    def format_number(self, value: float, zero_placeholder: str = "-") -> str:
        """
        Format a numeric value for display.
        
        Args:
            value: Number to format
            zero_placeholder: String to display for zero values (default "-")
            
        Returns:
            Formatted string (rounded integer or placeholder)
            
        Example:
            >>> format_number(1234.56)
            '1235'
            >>> format_number(0.0)
            '-'
        """
        if abs(value) >= 1:
            return f"{int(round(value))}"
        elif 0 < abs(value) < 1:
            return f"{value:.1f}"
        elif value == 0:
            return zero_placeholder
        else:
            return "0"
    
    def format_percentage(self, numerator: float, denominator: float, 
                         zero_placeholder: str = "-") -> str:
        """
        Format a percentage value for display.
        
        Args:
            numerator: Numerator value
            denominator: Denominator value
            zero_placeholder: String to display when denominator is zero
            
        Returns:
            Formatted percentage string like "12.5%" or placeholder
            
        Example:
            >>> format_percentage(125, 100)
            '25.0%'
            >>> format_percentage(50, 0)
            '-'
        """
        if denominator and denominator != 0:
            pct = (numerator / denominator * 100) - 100
            return f"{pct:.1f}%"
        return zero_placeholder
    
    def _calculate_pdf_layout(self, pdf_data: List[List], margin: float = 1 * inch) -> Tuple:
        """
        Calculate optimal page size and scaling factor for PDF table.
        
        Dynamically determines whether portrait or landscape orientation is needed,
        and calculates a scaling factor to ensure the table fits within the page width.
        
        Args:
            pdf_data: 2D list of table data (rows x columns)
            margin: Page margin in points (default 1 inch = 72 points)
            
        Returns:
            Tuple of (pagesize, col_widths, scale_factor)
            - pagesize: ReportLab page size tuple (width, height)
            - col_widths: List of column widths for the table
            - scale_factor: Factor to scale font sizes if needed
        """
        # Calculate approximate column widths based on content
        num_cols = len(pdf_data[0]) if pdf_data else 0
        if num_cols == 0:
            return A4, [], 1.0
        
        # Estimate column widths based on max content length
        col_widths = []
        for col_idx in range(num_cols):
            max_len = max(len(str(row[col_idx])) if col_idx < len(row) else 0 for row in pdf_data)
            # Approximate width: ~7 points per character for Helvetica 10pt, with padding
            col_width = max(max_len * 7, 50)  # Minimum 50 points per column
            col_widths.append(col_width)
        
        total_table_width = sum(col_widths)
        
        # Calculate available width for different page sizes
        a4_portrait_available = A4[0] - (2 * margin)  # ~446 points
        a4_landscape_available = A4[1] - (2 * margin)  # ~743 points
        letter_portrait_available = letter[0] - (2 * margin)  # ~468 points
        letter_landscape_available = letter[1] - (2 * margin)  # ~648 points
        
        # Choose page orientation and size based on content width
        if total_table_width <= a4_portrait_available:
            pagesize = A4
            available_width = a4_portrait_available
        elif total_table_width <= a4_landscape_available:
            pagesize = landscape(A4)
            available_width = a4_landscape_available
        else:
            # Use landscape A4 and scale content to fit
            pagesize = landscape(A4)
            available_width = a4_landscape_available
        
        # Calculate scaling factor to fit content
        if total_table_width > available_width:
            scale_factor = available_width / total_table_width
            # Scale column widths proportionally
            col_widths = [w * scale_factor for w in col_widths]
        else:
            scale_factor = 1.0
            # Distribute extra space proportionally
            extra_space = available_width - total_table_width
            per_col_extra = extra_space / num_cols
            col_widths = [w + per_col_extra for w in col_widths]
        
        return pagesize, col_widths, scale_factor
    
    def export_to_csv(self, df: pd.DataFrame, path: str, headers: List[str]) -> None:
        """
        Export DataFrame to CSV with custom headers.
        
        Args:
            df: DataFrame to export
            path: Output file path
            headers: Column headers to use
        """
        output_df = df.copy()
        if len(headers) == len(output_df.columns):
            output_df.columns = headers
        output_df.to_csv(path, index=False, sep=',')
        logging.info(f"Report exported to {path}")
    
    def export_to_txt(self, content: str, path: str) -> None:
        """
        Export formatted text content to file.
        
        Args:
            content: Text content to write
            path: Output file path
        """
        with open(path, 'w') as f:
            f.write(content)
        logging.info(f"Report exported to {path}")
    
    def export_to_html(self, df: pd.DataFrame, path: str, headers: List[str], 
                      title: str = "Report") -> None:
        """
        Export DataFrame to HTML table for Outlook.
        
        Args:
            df: DataFrame to export
            path: Output file path
            headers: Column headers to use
            title: HTML page title
        """
        html_content = f"""
        <html>
        <head><title>{title}</title></head>
        <body>
        <h2>{title}</h2>
        <table border="1" style="border-collapse: collapse; font-family: Arial, sans-serif; font-size: 12px;">
        <tr style="background-color: #f0f0f0;">
        """
        
        for header in headers:
            align = "left" if headers.index(header) == 0 else "right"
            html_content += f'<th style="padding: 8px; text-align: {align};">{header}</th>'
        
        html_content += "</tr>\n"
        
        for _, row in df.iterrows():
            is_total = row.get('is_total', False) or row.get('is_grand_total', False)
            bg_color = '#e6f3ff' if is_total else 'white'
            html_content += f'<tr style="background-color: {bg_color};">\n'
            
            for col in df.columns:
                value = row[col]
                align = "left" if df.columns.get_loc(col) == 0 else "right"
                html_content += f'<td style="padding: 8px; text-align: {align};">{value}</td>'
            
            html_content += "</tr>\n"
        
        html_content += "</table></body></html>"
        
        with open(path, 'w') as f:
            f.write(html_content)
        logging.info(f"Report exported to {path}")
    
    @abstractmethod
    def calculate_report(self) -> pd.DataFrame:
        """
        Calculate report data.
        
        Must be implemented by subclasses to generate report-specific
        calculations and return a DataFrame with the report data.
        
        Returns:
            DataFrame containing calculated report data
        """
        pass
    
    @abstractmethod
    def render_report(self, df: pd.DataFrame) -> None:
        """
        Render report to console.
        
        Must be implemented by subclasses to display report data
        in a formatted table to stdout.
        
        Args:
            df: DataFrame containing report data from calculate_report()
        """
        pass
    
    @abstractmethod
    def get_report_headers(self) -> List[str]:
        """
        Get column headers for the report.
        
        Must be implemented by subclasses to return the appropriate
        headers for CSV/PDF/HTML exports.
        
        Returns:
            List of column header strings
        """
        pass
    
    @abstractmethod
    def get_report_title(self) -> str:
        """
        Get the report title for exports.
        
        Must be implemented by subclasses to return the report title.
        
        Returns:
            Report title string
        """
        pass
    
    @abstractmethod
    def format_row_for_export(self, row: pd.Series) -> List[str]:
        """
        Format a DataFrame row for export.
        
        Must be implemented by subclasses to format row values
        according to report-specific requirements.
        
        Args:
            row: DataFrame row as Series
            
        Returns:
            List of formatted string values
        """
        pass
    
    def export_report(self, df: pd.DataFrame, base_path: str) -> None:
        """
        Export report to multiple formats (CSV, TXT, HTML, PDF, XLSX).
        
        Args:
            df: DataFrame containing report data from calculate_report()
            base_path: Base path for output files (e.g., 'report.csv')
                       Other formats will use same base with different extensions
        """
        now = self.now
        headers = self.get_report_headers()
        title = self.get_report_title()

        is_eom_anchor = (
            getattr(self, "_report_date", None) is not None
            and now.day == calendar.monthrange(now.year, now.month)[1]
        )
        period_label = "EOM" if is_eom_anchor else "MTD"
        if is_eom_anchor:
            date_range = f"{now.strftime('%B')} 1-{now.day}, {now.year}"
        else:
            date_range = format_mtd_date_range(now)
        
        # Build export rows once so styling metadata stays aligned with content.
        text_lines = []
        export_entries = []
        pdf_data = [headers]
        
        for _, row in df.iterrows():
            if row.get('is_spacer'):
                empty_row = [''] * len(headers)
                text_lines.append('')
                export_entries.append({
                    'row': row,
                    'formatted': empty_row,
                    'is_spacer': True,
                })
                pdf_data.append(empty_row)
                continue

            formatted = self.format_row_for_export(row)
            text_lines.append('\t'.join(formatted))
            export_entries.append({
                'row': row,
                'formatted': formatted,
                'is_spacer': False,
            })
            pdf_data.append(formatted)
        
        # Build text content with headers
        text_content = '\t'.join(headers) + '\n'
        text_content += '\n'.join(text_lines)
        
        # Build HTML content
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>{title}</title>
    <style>
        table {{ border-collapse: collapse; font-family: Arial, sans-serif; font-size: 12px; }}
        th {{ background-color: #4472C4; color: white; padding: 8px; text-align: center; }}
        td {{ padding: 8px; border: 1px solid #ddd; }}
        .total {{ background-color: #d9e8fb; font-weight: bold; }}
        .grand-total {{ background-color: #b8d4f1; font-weight: bold; }}
        .spacer td {{ border: none; padding: 6px; background: white; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
    </style>
</head>
<body>
<h2>{title} ({period_label}: {date_range})</h2>
<table>
<tr>"""
        
        for i, header in enumerate(headers):
            align = "left" if i == 0 else "right"
            html_content += f'<th style="text-align: {align};">{header}</th>'
        html_content += "</tr>\n"
        
        for entry in export_entries:
            row = entry['row']
            row_data = entry['formatted']

            if entry['is_spacer']:
                html_content += '<tr class="spacer">'
                for _ in headers:
                    html_content += '<td></td>'
                html_content += '</tr>\n'
                continue

            is_total = row.get('is_total', False)
            is_grand_total = row.get('is_grand_total', False)
            should_bold_override = row.get('should_bold', None)

            # Use should_bold override if present, otherwise fall back to is_total/is_grand_total.
            should_apply_bold = should_bold_override if should_bold_override is not None else (is_total or is_grand_total)

            if is_grand_total and should_apply_bold:
                row_class = 'grand-total'
            elif is_total and should_apply_bold:
                row_class = 'total'
            else:
                row_class = ''

            html_content += f'<tr class="{row_class}">'

            for i, val in enumerate(row_data):
                align = "left" if i == 0 else "right"
                weight = "font-weight:bold;" if should_apply_bold else ""
                html_content += f'<td style="text-align:{align};{weight}">{val}</td>'
            html_content += "</tr>\n"
        
        html_content += "</table></body></html>"
        
        # === Write CSV ===
        csv_path = base_path
        csv_df = df[~df.get('is_spacer', False)].copy() if 'is_spacer' in df.columns else df.copy()
        # Create export columns from formatted data
        export_rows = []
        for _, row in df.iterrows():
            if row.get('is_spacer'):
                continue
            export_rows.append(self.format_row_for_export(row))
        
        export_df = pd.DataFrame(export_rows, columns=headers)
        export_df.to_csv(csv_path, index=False, sep=',')
        print(f"Report exported to {csv_path}")
        
        # === Write TXT ===
        txt_path = base_path.replace('.csv', '.txt')
        with open(txt_path, 'w') as f:
            f.write(text_content)
        print(f"Report exported to {txt_path}")
        
        # === Write HTML ===
        html_path = base_path.replace('.csv', '.html')
        with open(html_path, 'w') as f:
            f.write(html_content)
        print(f"Report exported to {html_path} (Outlook-ready HTML table)")
        
        # === Write XLSX (styled Excel) ===
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            xlsx_path = base_path.replace('.csv', '.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            total_fill = PatternFill(start_color="D9E8FB", end_color="D9E8FB", fill_type="solid")
            grand_total_fill = PatternFill(start_color="B8D4F1", end_color="B8D4F1", fill_type="solid")
            total_font = Font(bold=True)
            grand_total_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_alignment = Alignment(horizontal='center', vertical='center')
            text_alignment = Alignment(horizontal='left', vertical='center')
            number_alignment = Alignment(horizontal='right', vertical='center')
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            row_idx = 2
            for _, row in df.iterrows():
                if row.get('is_spacer'):
                    row_idx += 1
                    continue
                
                formatted = self.format_row_for_export(row)
                
                for col_idx, val in enumerate(formatted, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = text_alignment if col_idx == 1 else number_alignment
                    cell.border = thin_border
                
                # Apply styling based on row type and explicit overrides
                is_total = row.get('is_total', False)
                is_grand_total = row.get('is_grand_total', False)
                should_bold_override = row.get('should_bold', None)
                
                # Use should_bold override if present, otherwise fall back to is_total/is_grand_total
                should_apply_bold = should_bold_override if should_bold_override is not None else (is_total or is_grand_total)
                
                if is_grand_total:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).font = grand_total_font
                        ws.cell(row=row_idx, column=col_idx).fill = grand_total_fill
                elif is_total:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).font = total_font
                        ws.cell(row=row_idx, column=col_idx).fill = total_fill
                elif should_apply_bold and not is_grand_total and not is_total:
                    # Apply bold only (no background color) if should_bold_override is True but not a total row
                    bold_font = Font(bold=True)
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).font = bold_font
                
                row_idx += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            for i in range(2, len(headers) + 1):
                ws.column_dimensions[chr(ord('A') + i - 1)].width = 12
            
            wb.save(xlsx_path)
            print(f"Report exported to {xlsx_path} (Excel format with formatting)")
            
        except ImportError:
            print("[WARNING] openpyxl not installed - skipping XLSX export")
        except Exception as e:
            print(f"[WARNING] Failed to create XLSX: {e}")
        
        # === Write PDF ===
        pdf_path = base_path.replace('.csv', '.pdf')
        
        # Define explicit margins (1 inch = 72 points)
        pdf_margin = 1 * inch
        
        # Calculate optimal page size and column widths dynamically
        pagesize, col_widths, scale_factor = self._calculate_pdf_layout(pdf_data, margin=pdf_margin)
        
        # Create document with explicit margins
        doc = SimpleDocTemplate(
            pdf_path, 
            pagesize=pagesize,
            leftMargin=pdf_margin,
            rightMargin=pdf_margin,
            topMargin=pdf_margin,
            bottomMargin=pdf_margin
        )
        styles = getSampleStyleSheet()
        
        pdf_title = Paragraph(f"{title} ({period_label}: {date_range})", styles['Heading1'])
        
        # Create table with calculated column widths
        table = Table(pdf_data, colWidths=col_widths)
        
        # Calculate font sizes based on scale factor (minimum 8pt)
        header_font_size = max(int(14 * scale_factor), 10)
        data_font_size = max(int(10 * scale_factor), 8)
        padding = max(int(12 * scale_factor), 6)
        
        # Style the table with scaled font sizes
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Left align first column
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), padding),
            ('TOPPADDING', (0, 0), (-1, -1), padding // 2),
            ('BOTTOMPADDING', (0, 1), (-1, -1), padding // 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        
        # Add special styling for totals
        row_idx = 1
        for _, row in df.iterrows():
            if row.get('is_spacer'):
                row_idx += 1
                continue
            is_total = row.get('is_total', False)
            is_grand_total = row.get('is_grand_total', False)
            should_bold_override = row.get('should_bold', None)
            
            # Use should_bold override if present, otherwise fall back to is_total/is_grand_total
            should_apply_bold = should_bold_override if should_bold_override is not None else (is_total or is_grand_total)
            
            if should_apply_bold:
                style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightblue)
                style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
            row_idx += 1
        
        table.setStyle(style)
        
        # Build PDF
        elements = [pdf_title, Spacer(1, 20), table]
        doc.build(elements)
        print(f"Report exported to {pdf_path} (PDF format, {pagesize[0]:.0f}x{pagesize[1]:.0f} pts)")
