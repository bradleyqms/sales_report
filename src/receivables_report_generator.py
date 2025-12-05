import pandas as pd
import json
import datetime
import os
import tempfile
import sys
import time
import logging
import warnings
from pathlib import Path
from dotenv import load_dotenv

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from sharepoint_client import SharePointHandler, download_inputs, upload_outputs
from qry_data_ingestion import process_qry_files
from qry_data_mapping import apply_mappings
from utils import print_progress, get_current_year, get_prior_year, get_current_month, format_mtd_date_range

class ManagementReportGenerator:
    def __init__(self, config_path, sales_path, budget_path, prior_path):
        self.config = self._load_config(config_path)
        try:
            self.df = pd.read_csv(sales_path)
            self.budget_df = pd.read_csv(budget_path)
            self.prior_df = pd.read_csv(prior_path)
        except FileNotFoundError as e:
            logging.error(f"Required data file not found: {e}")
            raise
        except pd.errors.EmptyDataError as e:
            logging.error(f"Data file is empty: {e}")
            raise
        
        self._prepare_data()
        
    def _load_config(self, path):
        try:
            with open(path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logging.error(f"Config file not found: {path}")
            raise
        except json.JSONDecodeError as e:
            logging.error(f"Invalid JSON in config file: {e}")
            raise
            
    def _prepare_data(self):
        # Dates
        now = datetime.datetime.now()
        self.current_month = now.month
        self.current_year = 2025
        self.prior_year = 2024
        
        # Filter Sales to AR (for QRY data, Document Type is 'AR', not 'AR Invoice')
        self.df = self.df[self.df['Document Type'] == 'AR'].copy()
        
        # Convert Sales to kEUR
        value_col = 'Value_in_EUR_converted' if 'Value_in_EUR_converted' in self.df.columns else 'Total Value (EUR)'
        self.df['kEUR'] = self.df[value_col].fillna(0) / 1000
        
        # Filter Budget for Current Month
        # Budget Date is DD/MM/YYYY
        self.budget_df['Date'] = pd.to_datetime(self.budget_df['Date'], format='%d/%m/%Y')
        self.budget_month = self.budget_df[self.budget_df['Date'].dt.month == self.current_month].copy()
        
        # Filter Prior for Same Month Last Year
        target_prior_date = f"{self.prior_year}-{self.current_month:02d}"
        self.prior_month = self.prior_df[self.prior_df['Date'].astype(str).str.startswith(target_prior_date)].copy()
        
    def calculate_report(self):
        report_data = []
        section_totals = {}
        grand_total = {'Sales': 0, 'Budget': 0, 'Prior': 0}
        
        for section in self.config['sections']:
            if section.get('is_grand_total'):
                # Before outputting grand total, subtract any company-level totals
                # so that grand total reflects base items only.
                deduction_sales = 0
                deduction_budget = 0
                deduction_prior = 0
                for key,vals in section_totals.items():
                    # Match titles like 'Company 1 Sales', 'Company 2 Sales', etc.
                    if isinstance(key, str) and key.startswith('Company ') and key.endswith(' Sales'):
                        deduction_sales += vals.get('sales', 0)
                        deduction_budget += vals.get('budget', 0)
                        deduction_prior += vals.get('prior', 0)

                adj_sales = grand_total['Sales'] - deduction_sales
                adj_budget = grand_total['Budget'] - deduction_budget
                adj_prior = grand_total['Prior'] - deduction_prior

                report_data.append({
                    'label': section['title'],
                    'sales': adj_sales,
                    'budget': adj_budget,
                    'prior': adj_prior,
                    'is_total': True,
                    'is_grand_total': True,
                    'is_spacer': False
                })
                continue
                
            if section.get('is_unmapped'):
                continue
                
            if section.get('is_total'):
                # Sum of other sections (e.g. Company 1 Total)
                t_sales = t_budget = t_prior = 0
                for comp in section['items'] if 'items' in section else section.get('components', []):
                    if comp in section_totals:
                        t_sales += section_totals[comp]['sales']
                        t_budget += section_totals[comp]['budget']
                        t_prior += section_totals[comp]['prior']
                
                report_data.append({
                    'label': section['title'],
                    'sales': t_sales,
                    'budget': t_budget,
                    'prior': t_prior,
                    'is_total': True,
                    'is_spacer': False,
                    'is_grand_total': False
                })
                
                # Add to grand total
                grand_total['Sales'] += t_sales
                grand_total['Budget'] += t_budget
                grand_total['Prior'] += t_prior
                continue

            # Regular Section
            sec_sales = 0
            sec_budget = 0
            sec_prior = 0
            
            # Get Section Totals first (for fallback calculation)
            # Filter by Company Group and Market Group
            c_group = section.get('company_group')
            m_group = section.get('market_group')
            
            # Base filters for the whole section
            sales_mask = (self.df['Company_Group'] == c_group)
            budget_mask = (self.budget_month['Company_Group'] == c_group)
            prior_mask = (self.prior_month['Company_Group'] == c_group)
            
            if m_group:
                sales_mask &= (self.df['Market_Group'] == m_group)
                budget_mask &= (self.budget_month['Market_Group'] == m_group)
                prior_mask &= (self.prior_month['Market_Group'] == m_group)
            
            section_total_sales = self.df[sales_mask]['kEUR'].sum()
            
            # Budget values already in kEUR/kUSD format
            if m_group == 'USA' and 'Value_kUSD' in self.budget_month.columns:
                section_total_budget = self.budget_month[budget_mask]['Value_kUSD'].sum()
            else:
                section_total_budget = self.budget_month[budget_mask]['Value_kEUR'].sum()
            
            section_total_prior = self.prior_month[prior_mask]['Value_kEUR'].sum()
            
            # Track allocated amounts to calculate fallback
            allocated_sales = 0
            allocated_budget = 0
            allocated_prior = 0
            
            rows = []
            
            for item in section.get('items', []):
                label = item['label']
                is_fallback = item.get('is_fallback', False)
                
                if is_fallback:
                    # Will calculate at end of loop
                    rows.append({'label': label, 'type': 'fallback'})
                    continue
                
                # Item specific filters
                filter_val = item.get('filter_value')
                filter_type = section.get('type') # 'region' or 'channel'
                
                # Sales Filter
                s_mask = sales_mask.copy()
                if filter_type == 'region':
                    s_mask &= (self.df['Region'] == filter_val)
                elif filter_type == 'channel':
                    s_mask &= (self.df['Channel_Level'] == filter_val)
                
                val_sales = self.df[s_mask]['kEUR'].sum()
                
                # Budget/Prior Filter
                # Check for override map (e.g. Company 3 channels mapping to regions)
                b_filter_val = item.get('budget_region_map', filter_val)
                
                b_mask = budget_mask.copy()
                p_mask = prior_mask.copy()
                
                lookup_col = 'Region' if 'budget_region_map' in item else ('Region' if filter_type == 'region' else 'Channel_Level')
                
                b_mask &= (self.budget_month[lookup_col] == b_filter_val)
                p_mask &= (self.prior_month[lookup_col] == b_filter_val)
                
                # Budget values are already in kEUR/kUSD format in source file
                if m_group == 'USA' and 'Value_kUSD' in self.budget_month.columns:
                    val_budget = self.budget_month[b_mask]['Value_kUSD'].sum()
                else:
                    val_budget = self.budget_month[b_mask]['Value_kEUR'].sum()
                
                val_prior = self.prior_month[p_mask]['Value_kEUR'].sum()
                
                rows.append({
                    'label': label,
                    'sales': val_sales,
                    'budget': val_budget,
                    'prior': val_prior,
                    'is_total': False,
                    'is_spacer': False
                })
                
                allocated_sales += val_sales
                allocated_budget += val_budget
                allocated_prior += val_prior
            
            # Process Fallback
            for i, row in enumerate(rows):
                if row.get('type') == 'fallback':
                    rem_sales = section_total_sales - allocated_sales
                    rem_budget = section_total_budget - allocated_budget
                    rem_prior = section_total_prior - allocated_prior
                    
                    # Only show if there is value
                    if abs(rem_sales) > 0.1 or abs(rem_budget) > 0.1 or abs(rem_prior) > 0.1:
                        rows[i] = {
                            'label': row['label'],
                            'sales': rem_sales,
                            'budget': rem_budget,
                            'prior': rem_prior,
                            'is_total': False,
                            'is_spacer': False
                        }
                    else:
                        rows[i] = None # Mark for removal
            
            # Add rows to report
            rows = [r for r in rows if r is not None]
            report_data.extend(rows)
            
            # Add Section Total if requested or if it's a component
            if section.get('show_total') or section.get('title') in ['Core Markets', 'UK', 'USA', 'Export']:
                report_data.append({
                    'label': section['title'], # or "Total " + section['title']
                    'sales': section_total_sales,
                    'budget': section_total_budget,
                    'prior': section_total_prior,
                    'is_total': True
                })
                
            # Store for aggregation
            section_totals[section['title']] = {
                'sales': section_total_sales,
                'budget': section_total_budget,
                'prior': section_total_prior
            }
            
            # Add spacer with consistent schema
            report_data.append({
                'label': '',
                'sales': 0.0,
                'budget': 0.0,
                'prior': 0.0,
                'is_spacer': True,
                'is_total': False,
                'is_grand_total': False
            })
            
            # Add to grand total for company sales sections
            if 'Sales' in section['title']:
                grand_total['Sales'] += section_total_sales
                grand_total['Budget'] += section_total_budget
                grand_total['Prior'] += section_total_prior

        # Create DataFrame with explicit type enforcement
        df = pd.DataFrame(report_data)
        
        # Convert types directly (no fillna needed since all fields are explicitly set)
        df['is_spacer'] = df['is_spacer'].astype(bool)
        df['is_total'] = df['is_total'].astype(bool)
        df['is_grand_total'] = df['is_grand_total'].astype(bool)
        df['sales'] = df['sales'].astype(float)
        df['budget'] = df['budget'].astype(float)
        df['prior'] = df['prior'].astype(float)
        df['label'] = df['label'].astype(str)
        
        return df

    def render_report(self, df):
        # Print Header
        now = datetime.datetime.now()
        month_name = now.strftime('%b')
        year_short = str(now.year)[2:]
        col_curr = f"{month_name}-{year_short}A MTD"
        
        print(f"{'kEUR':<30} {col_curr:>15} {'Budget':>10} {'Prior':>10} {'% vs Bud':>10}")
        print("-" * 75)
        
        for _, row in df.iterrows():
            if 'is_spacer' in df.columns and row.get('is_spacer') == True:
                print()
                continue
                
            label = row['label']
            sales = row['sales']
            budget = row['budget']
            prior = row['prior']
            
            # Add extra space above Company Sales totals
            if row.get('is_total') and 'Sales' in label:
                print()
            
            # Calculate %
            pct = (sales / budget * 100) if budget and budget != 0 else 0
            
            # Format values (already in kEUR)
            s_str = f"{int(round(sales))}" if abs(sales) >= 0.5 else ("-" if sales == 0 else "0")
            b_str = f"{int(round(budget))}" if abs(budget) >= 0.5 else ("-" if budget == 0 else "0")
            p_str = f"{int(round(prior))}" if abs(prior) >= 0.5 else ("-" if prior == 0 else "0")
            pct_str = f"{pct:.1f}%" if budget and budget != 0 else "-"
            
            print(f"{label:<30} {s_str:>10} {b_str:>10} {p_str:>10} {pct_str:>10}")
            
            if row.get('is_total') or row.get('is_grand_total'):
                print("-" * 75)
    
    def export_report(self, df, base_path):
        """Export the report in formatted text style to CSV/TXT, HTML for Outlook, and PDF."""
        # Define column widths for text format
        now = datetime.datetime.now()
        month_name = now.strftime('%b')
        year_short = str(now.year)[2:]
        col_curr = f"{month_name}-{year_short}A MTD"
        col_widths = [35, 15, 12, 12, 12]
        headers = ['kEUR', col_curr, 'Budget', 'Prior', '% vs Bud']
        
        # Create text format
        header_line = ''.join(f"{h:<{w}}" for h, w in zip(headers, col_widths))
        separator = '-' * len(header_line)
        
        formatted_lines = [header_line, separator]
        
        for _, row in df.iterrows():
            if 'is_spacer' in df.columns and row.get('is_spacer') == True:
                formatted_lines.append('')
                continue
                
            label = row['label']
            sales = row['sales']
            budget = row['budget']
            prior = row['prior']
            
            pct = (sales / budget * 100) if budget and budget != 0 else 0
            
            s_str = f"{int(round(sales))}" if abs(sales) >= 0.5 else ("-" if sales == 0 else "0")
            b_str = f"{int(round(budget))}" if abs(budget) >= 0.5 else ("-" if budget == 0 else "0")
            p_str = f"{int(round(prior))}" if abs(prior) >= 0.5 else ("-" if prior == 0 else "0")
            pct_str = f"{pct:.1f}%" if budget and budget != 0 else "-"
            
            row_line = f"{label:<{col_widths[0]}}{s_str:>{col_widths[1]}}{b_str:>{col_widths[2]}}{p_str:>{col_widths[3]}}{pct_str:>{col_widths[4]}}"
            formatted_lines.append(row_line)
            
            if row.get('is_total') or row.get('is_grand_total'):
                formatted_lines.append(separator)
        
        text_content = '\n'.join(formatted_lines)
        
        # Create HTML format for Outlook
        html_content = f"""
        <html>
        <body>
        <table border="1" style="border-collapse: collapse; font-family: Arial, sans-serif; font-size: 12px;">
        <tr style="background-color: #f0f0f0;">
            <th style="padding: 8px; text-align: left;">{headers[0]}</th>
            <th style="padding: 8px; text-align: right;">{headers[1]}</th>
            <th style="padding: 8px; text-align: right;">{headers[2]}</th>
            <th style="padding: 8px; text-align: right;">{headers[3]}</th>
            <th style="padding: 8px; text-align: right;">{headers[4]}</th>
        </tr>
        """
        
        for _, row in df.iterrows():
            if 'is_spacer' in df.columns and row.get('is_spacer') == True:
                html_content += '<tr><td colspan="5" style="height: 10px;"></td></tr>\n'
                continue
                
            label = row['label']
            sales = row['sales']
            budget = row['budget']
            prior = row['prior']
            
            pct = (sales / budget * 100) if budget and budget != 0 else 0
            
            s_str = f"{int(round(sales))}" if abs(sales) >= 0.5 else ("-" if sales == 0 else "0")
            b_str = f"{int(round(budget))}" if abs(budget) >= 0.5 else ("-" if budget == 0 else "0")
            p_str = f"{int(round(prior))}" if abs(prior) >= 0.5 else ("-" if prior == 0 else "0")
            pct_str = f"{pct:.1f}%" if budget and budget != 0 else "-"
            
            # Highlight totals
            bg_color = '#e6f3ff' if row.get('is_total') or row.get('is_grand_total') else 'white'
            
            html_content += f"""
            <tr style="background-color: {bg_color};">
                <td style="padding: 8px;">{label}</td>
                <td style="padding: 8px; text-align: right;">{s_str}</td>
                <td style="padding: 8px; text-align: right;">{b_str}</td>
                <td style="padding: 8px; text-align: right;">{p_str}</td>
                <td style="padding: 8px; text-align: right;">{pct_str}</td>
            </tr>
            """
        
        html_content += "</table></body></html>"
        
        # Create proper CSV format with comma separators
        csv_df = df.copy()
        # Filter out spacer rows for CSV (is_spacer is already bool from calculate_report)
        if 'is_spacer' in csv_df.columns:
            csv_df = csv_df[~csv_df['is_spacer']]
        csv_df['% vs Bud'] = csv_df.apply(lambda row: f"{(row['sales'] / row['budget'] * 100):.1f}%" if row['budget'] and row['budget'] != 0 else "-", axis=1)
        csv_df[col_curr] = csv_df['sales'].apply(lambda x: f"{int(round(x))}" if abs(x) >= 0.5 else ("-" if x == 0 else "0"))
        csv_df['Budget'] = csv_df.apply(lambda row: f"{int(round(row['budget']/1000))}" if row.get('budget_needs_division', True) and abs(row['budget']) >= 500 else (f"{int(round(row['budget']))}" if not row.get('budget_needs_division', True) and abs(row['budget']) >= 0.5 else ("-" if row['budget'] == 0 else "0")), axis=1)
        csv_df['Prior'] = csv_df['prior'].apply(lambda x: f"{int(round(x))}" if abs(x) >= 0.5 else ("-" if x == 0 else "0"))
        csv_df = csv_df.rename(columns={'label': 'kEUR'})
        csv_df = csv_df[['kEUR', col_curr, 'Budget', 'Prior', '% vs Bud']]
        
        # Write to CSV file (proper CSV format with commas)
        csv_path = base_path
        csv_df.to_csv(csv_path, index=False, sep=',')
        print(f"Report exported to {csv_path}")
        
        # Write to XLSX file with formatting
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            xlsx_path = base_path.replace('.csv', '.xlsx')
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Management Report"
            
            # Define styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            total_font = Font(name='Calibri', size=11, bold=True)
            total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            grand_total_font = Font(name='Calibri', size=12, bold=True)
            grand_total_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
            
            number_alignment = Alignment(horizontal='right', vertical='center')
            text_alignment = Alignment(horizontal='left', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Write headers (reuse headers from above with MTD)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            row_idx = 2
            for _, row in df.iterrows():
                if row.get('is_spacer'):
                    row_idx += 1
                    continue
                
                # Get values
                label = row['label']
                sales = row['sales']
                budget = row['budget']
                prior = row['prior']
                pct = (sales / budget * 100) if budget and budget != 0 else 0
                
                # Format values
                s_val = int(round(sales)) if abs(sales) >= 0.5 else None
                b_val = int(round(budget)) if abs(budget) >= 0.5 else None
                p_val = int(round(prior)) if abs(prior) >= 0.5 else None
                pct_val = f"{pct:.1f}%" if budget and budget != 0 else "-"
                
                # Write row
                ws.cell(row=row_idx, column=1, value=label).alignment = text_alignment
                ws.cell(row=row_idx, column=2, value=s_val if s_val else "-").alignment = number_alignment
                ws.cell(row=row_idx, column=3, value=b_val if b_val else "-").alignment = number_alignment
                ws.cell(row=row_idx, column=4, value=p_val if p_val else "-").alignment = number_alignment
                ws.cell(row=row_idx, column=5, value=pct_val).alignment = number_alignment
                
                # Apply styling based on row type
                is_total = row.get('is_total', False)
                is_grand_total = row.get('is_grand_total', False)
                
                if is_grand_total:
                    for col_idx in range(1, 6):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = grand_total_font
                        cell.fill = grand_total_fill
                        cell.border = thin_border
                elif is_total:
                    for col_idx in range(1, 6):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = total_font
                        cell.fill = total_fill
                        cell.border = thin_border
                else:
                    for col_idx in range(1, 6):
                        ws.cell(row=row_idx, column=col_idx).border = thin_border
                
                row_idx += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            
            # Save workbook
            wb.save(xlsx_path)
            print(f"Report exported to {xlsx_path} (Excel format with formatting)")
            
        except ImportError:
            print("[WARNING] openpyxl not installed - skipping XLSX export")
        except Exception as e:
            print(f"[WARNING] Failed to create XLSX: {e}")
        
        # Write to TXT file (text format)
        txt_path = base_path.replace('.csv', '.txt')
        with open(txt_path, 'w') as f:
            f.write(text_content)
        print(f"Report exported to {txt_path}")
        
        # Write to HTML file (for Outlook)
        html_path = base_path.replace('.csv', '.html')
        with open(html_path, 'w') as f:
            f.write(html_content)
        print(f"Report exported to {html_path} (Outlook-ready HTML table)")
        
        # Create PDF format
        pdf_path = base_path.replace('.csv', '.pdf')
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # PDF title with MTD date range
        date_range = now.strftime('%B 1-%d, %Y')
        title = Paragraph(f"QRY Management Report (MTD: {date_range})", styles['Heading1'])
        
        # Prepare table data
        pdf_data = [headers]
        
        for _, row in df.iterrows():
            if 'is_spacer' in df.columns and row.get('is_spacer') == True:
                pdf_data.append(['', '', '', '', ''])  # Empty row for spacing
                continue
                
            label = row['label']
            sales = row['sales']
            budget = row['budget']
            prior = row['prior']
            
            pct = (sales / budget * 100) if budget and budget != 0 else 0
            
            s_str = f"{int(round(sales))}" if abs(sales) >= 0.5 else ("-" if sales == 0 else "0")
            b_str = f"{int(round(budget))}" if abs(budget) >= 0.5 else ("-" if budget == 0 else "0")
            p_str = f"{int(round(prior))}" if abs(prior) >= 0.5 else ("-" if prior == 0 else "0")
            pct_str = f"{pct:.1f}%" if budget and budget != 0 else "-"
            
            pdf_data.append([label, s_str, b_str, p_str, pct_str])
        
        # Create table
        table = Table(pdf_data)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Left align first column
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        
        # Add special styling for totals
        row_idx = 1
        for _, row in df.iterrows():
            if row.get('is_total') or row.get('is_grand_total'):
                style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightblue)
                style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
            row_idx += 1
        
        table.setStyle(style)
        
        # Build PDF
        elements = [title, Spacer(1, 20), table]
        doc.build(elements)
        print(f"Report exported to {pdf_path} (PDF format)")

if __name__ == "__main__":
    start_time = datetime.datetime.now()
    # Load environment variables
    load_dotenv()

    # SharePoint configuration (use environment variables for security)
    SHAREPOINT_SITE_URL = os.getenv('SHAREPOINT_SITE_URL')
    CLIENT_ID = os.getenv('SHAREPOINT_CLIENT_ID')
    CLIENT_SECRET = os.getenv('SHAREPOINT_CLIENT_SECRET')
    
    project_root = Path(__file__).parent.parent
    
    use_sharepoint = all([SHAREPOINT_SITE_URL, CLIENT_ID, CLIENT_SECRET])
    
    if use_sharepoint:
        print("Starting QRY Report Generation with SharePoint Data")
        print("=" * 60)
        
        # Initialize progress
        total_steps = 5
        current_step = 0
        
        # Initialize SharePoint handler (suppress connection message)
        original_stdout = sys.stdout
        sys.stdout = open(os.devnull, 'w')
        try:
            sp_handler = SharePointHandler(SHAREPOINT_SITE_URL, CLIENT_ID, CLIENT_SECRET, quiet=True)
        finally:
            sys.stdout = original_stdout
        
        # Create temp directory for downloads
        with tempfile.TemporaryDirectory() as temp_dir:
            current_step += 1
            print_progress(current_step, total_steps, "Downloading QRY files from SharePoint...")
            
            # List of QRY files to download
            qry_files = [
                "QRY_AR_MTD_CH.csv", "QRY_AR_MTD_Export.csv", "QRY_AR_MTD_Gmbh.csv", 
                "QRY_AR_MTD_UK.csv", "QRY_AR_MTD_USA.csv", 
                "QRY_CN_MTD_CH.csv", "QRY_CN_MTD_GmbH.csv", "QRY_CN_MTD_GmbH1.csv", 
                "QRY_CN_MTD_UK.csv", "QRY_CN_MTD_USA.csv", 
                "QRY_SO_OPEN_MTD_CH.csv", "QRY_SO_OPEN_MTD_Gmbh.csv", "QRY_SO_OPEN_MTD_USA.csv", 
                "QRY_SO_TOTAL_MTD_CH.csv", "QRY_SO_TOTAL_MTD_Gmbh.csv", "QRY_SO_TOTAL_MTD_USA.csv"
            ]
            
            sp_base_path = "/sites/DATAANDREPORTING/Shared Documents/SAP Extracts/"
            
            # Download QRY files (suppress individual prints)
            downloaded_count = 0
            original_stdout = sys.stdout
            sys.stdout = open(os.devnull, 'w')  # Suppress prints during downloads
            try:
                for filename in qry_files:
                    sp_path = sp_base_path + filename
                    local_path = os.path.join(temp_dir, filename)
                    try:
                        sp_handler.download_file(sp_path, local_path)
                        downloaded_count += 1
                    except Exception as e:
                        pass  # Silent failure for individual files
            finally:
                sys.stdout = original_stdout  # Restore stdout
            
            print()  # Move to new line after progress bar
            print(f"[OK] Downloaded {downloaded_count} QRY files from SharePoint")
            
            current_step += 1
            print_progress(current_step, total_steps, "Processing QRY data...")
            qry_df = process_qry_files(temp_dir)
            
            current_step += 1
            print_progress(current_step, total_steps, "Downloading support files...")
            
            # Define other SharePoint paths
            other_paths = {
                'mapping': '/sites/DATAANDREPORTING/Shared Documents/SAP Extracts/entity_mappings.csv',
                'budget': '/sites/DATAANDREPORTING/Shared Documents/SAP Extracts/budget_2025_processed.csv',
                'prior': '/sites/DATAANDREPORTING/Shared Documents/SAP Extracts/prior_sales_2024_processed.csv'
            }
            
            local_paths = {}
            original_stdout = sys.stdout
            sys.stdout = open(os.devnull, 'w')  # Suppress prints during downloads
            try:
                for key, sp_path in other_paths.items():
                    local_path = os.path.join(temp_dir, os.path.basename(sp_path))
                    try:
                        sp_handler.download_file(sp_path, local_path)
                        local_paths[key] = local_path
                    except Exception as e:
                        # Fallback to local paths
                        if key == 'mapping':
                            local_paths[key] = str(project_root / 'data/inputs/mappings/entity_mappings.csv')
                        elif key == 'budget':
                            local_paths[key] = str(project_root / 'data/inputs/budget/budget_2025_processed.csv')
                        elif key == 'prior':
                            local_paths[key] = str(project_root / 'data/inputs/prior_years/prior_sales_2024_processed.csv')
            finally:
                sys.stdout = original_stdout  # Restore stdout
            
            current_step += 1
            print_progress(current_step, total_steps, "Applying entity mappings...")
            mapping_df = pd.read_csv(local_paths['mapping'])
            mapped_df = apply_mappings(qry_df, mapping_df)
            
            current_step += 1
            print_progress(current_step, total_steps, "Generating management report...")
            
            # Save mapped data locally for reference/debugging
            mapped_path = os.path.join(temp_dir, 'qry_unified_mapped_2025.csv')
            mapped_df.to_csv(mapped_path, index=False)
            
            # Run the report generator with processed data
            generator = ManagementReportGenerator(
                str(project_root / 'src/config/report_structure.json'),
                mapped_path,
                local_paths['budget'],
                local_paths['prior']
            )
            df = generator.calculate_report()
            print()  # Move to new line before report output
            generator.render_report(df)
            
            # Generate timestamped filename
            now = datetime.datetime.now()
            timestamp = now.strftime('%Y%m%d_%H%M%S')
            base_filename = f'management_report_qry_2025_{timestamp}'
            
            # Export to data/outputs folder
            output_dir = str(project_root / 'data/outputs')
            os.makedirs(output_dir, exist_ok=True)
            local_base_path = os.path.join(output_dir, base_filename)
            generator.export_report(df, local_base_path + '.csv')
            
            print(f"\n[SUCCESS] Report generation complete! Files saved to {output_dir}/")
            print(f"[INFO] Exported in 4 formats: CSV, TXT, HTML, and PDF")
            print(f"[INFO] Processed {len(qry_df)} QRY records -> {len(mapped_df)} mapped records")
            print("=" * 60)
    else:
        print("SharePoint credentials not found. Using local files.")
        # Fallback to local file processing
        project_root = Path(__file__).parent.parent
        generator = ManagementReportGenerator(
            project_root / 'src/config/report_structure.json',
            project_root / 'data/outputs/qry_unified_mapped_2025.csv',
            project_root / 'data/inputs/budget/budget_2025_processed.csv',
            project_root / 'data/inputs/prior_years/prior_sales_2024_processed.csv'
        )
        df = generator.calculate_report()
        generator.render_report(df)
        
        # Generate timestamped filename
        now = datetime.datetime.now()
        timestamp = now.strftime('%Y%m%d_%H%M%S')
        output_path = project_root / f'data/outputs/management_report_qry_2025_{timestamp}.csv'
        
        generator.export_report(df, output_path)
    
    end_time = datetime.datetime.now()
    print("\nGenerator runtime: {:.2f} seconds".format((end_time - start_time).total_seconds()))